import os
import sys
import openpyxl
from datetime import datetime


USER_DB = "db.xlsx"
INV_DB = "inventory.xlsx"

def clear_screen():
    os.system("cls")

def show_welcome_screen():
    print("\n========================================")
    print("     INVENTORY MANAGEMENT SYSTEM")
    print("========================================\n\n\n")

def login():
    clear_screen()
    show_welcome_screen()
    wb = openpyxl.load_workbook(USER_DB)
    sheet = wb["Users"]
    username = input("Enter username: ")
    password = input("Enter password: ")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == username and row[2] == password:
            print("Login successful!\n")
            return row[3]

    print("Invalid credentials")
    input()
    return None

def init_inventory():
    try:
        openpyxl.load_workbook(INV_DB)
    except:
        wb = openpyxl.Workbook()

        product_sheet = wb.active
        product_sheet.title = "Products"
        product_sheet.append(["ID", "Name", "Price", "Quantity","GST"])

        trans_sheet = wb.create_sheet("Transactions")
        trans_sheet.append(["Date", "Product","Price" ,"Quantity","GST" ,"Total"])

        wb.save(INV_DB)


def generate_product_id():

    wt = openpyxl.load_workbook(INV_DB)
    sheet = wt["Products"]

    
    ids=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            ids.append(int(row[0]))

    if ids:
        new_id = max(ids) + 1
    else:
        new_id = 1001

    return new_id

def add_product():
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Products"]

    pid=generate_product_id()
    print("Generated Product ID:", pid)
    name = input("Product Name: ")
    price = float(input("Price: "))
    qty = int(input("Quantity: "))
    gst=float(input("Enter GST: "))
    sheet.append([pid, name, price, qty,gst])
    wb.save(INV_DB)

    print("Product added successfully")


def view_products():
    clear_screen()
    show_welcome_screen()
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Products"]

    print("--- Product List ---\n")
    print(f"{'ID':<10}{'Name':<25}{'Price':<15}{'Qty':<10}{'GST':<10}")
    print("-" * 80)
    data_prsnt=False
    for row in sheet.iter_rows(min_row=2,values_only=True):
        pid=row[0]
        name =row[1]
        price = row[2]
        qty =row[3]
        gst=row[4]
        data_prsnt=True
        print(f"{pid:<10}{name:<25}{price:<15}{qty:<10}{gst:<10}")
    if data_prsnt==False:
        print("No products available.")

def update_product():
    clear_screen()
    show_welcome_screen()
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Products"]
    pid = int(input("Enter Product ID to update: "))

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == pid:
            print("\n--- Current Details ---\n")
            print("Name: ", row[1].value)
            print("Price: ", row[2].value)
            print("Quantity: ", row[3].value)
            print("GST: ", row[4].value)
            print("\n\n")
            row[1].value = input("New Name: ")
            row[2].value = float(input("New Price: "))
            row[3].value = int(input("New Quantity: "))
            row[4].value = int(input("New GST: "))
            wb.save(INV_DB)
            print("Product updated")
            return

    print("Product not found")

def delete_product():
    clear_screen()
    show_welcome_screen()
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Products"]

    pid = int(input("Enter Product ID to delete: "))

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):

        if row[0].value == pid:
            print("\n--- Details ---\n")
            print("Name: ", row[1].value)
            print("Price: ", row[2].value)
            print("Quantity: ", row[3].value)
            print("GST: ", row[4].value)
            print("\n\n")
            x=input("Are You Sure? You Want To Delete This Item? (y/n): ")
            if x=="y":
                sheet.delete_rows(i)
                wb.save(INV_DB)
                print("\nProduct deleted")
            else:
                print("\nItem Not Deleted!")
            return

    print("Product not found")


def sell_product():
    
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Products"]
    trans = wb["Transactions"]
    try:
        openpyxl.load_workbook("crnt_bill.xlsx")
    except:
        wq = openpyxl.Workbook()

        product_sheet = wq.active
        product_sheet.title = "Products"
        product_sheet.append(["Name", "Price", "Quantity","GST","Total"])

        wq.save("crnt_bill.xlsx")
    wo=openpyxl.load_workbook("crnt_bill.xlsx")
    sheet_wo = wo["Products"]
    checker=False
    
    while(1):
        clear_screen()
        show_welcome_screen()
        checker_2=False
        print("--- Sell Product ---\n")
        pid = int(input("Enter Product ID: "))

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == pid:
                print("\n--- Details ---\n")
                print("Name: ", row[1].value)
                print("Price: ", row[2].value)
                print("Quantity: ", row[3].value)
                print("\n\n")
                qty = int(input("Quantity to sell: "))
                if row[3].value >= qty:
                    total = row[2].value * qty
                    row[3].value -= qty
                    gst_amt=(total+row[4].value)/100
                    total=total+gst_amt
                    trans.append([
                        datetime.now().strftime("%Y-%m-%d %H:%M"),
                        row[1].value,
                        row[2].value,
                        qty,
                        row[4].value,
                        total
                    ])
                    sheet_wo.append([row[1].value,row[2].value,row[3].value,row[4].value,total])
                    checker=True
                    checker_2=True
                    print("ADD Successfully!\n")
                    wb.save(INV_DB)
                    wo.save("crnt_bill.xlsx")
                    break

                else:
                    print("Not enough stock")
                    break
        if checker_2==False:
            print("Product not found")
        x=input("\nWant to Search More? (y/n): ")
        if x!="y":
            break
    if checker:
        wb = openpyxl.load_workbook("crnt_bill.xlsx")
        sheet = wb["Products"]

        print("\n--- Product List ---\n")
        print(f"{'Name':<25}{'Price':<15}{'Qty':<10}{'GST':<10}{'Total':<15}")
        print("-" * 80)
        grand_total=0
        data_prsnt=False
        for row in sheet.iter_rows(min_row=2,values_only=True):
            name =row[0]
            price = row[1]
            qty =row[2]
            gst=row[3]
            total=row[4]
            grand_total+=total
            data_prsnt=True
            print(f"{name:<25}{price:<15}{qty:<10}{gst:<10}{total:<15}")
        if data_prsnt==False:
            print("\nNo products available.")
        else:
            print("\nTotal Bill: ",grand_total)
            print("\nProduct sold successfully")
    os.remove("crnt_bill.xlsx")
    return       
        

        

def view_transactions():
    clear_screen()
    show_welcome_screen()
    wb = openpyxl.load_workbook(INV_DB)
    sheet = wb["Transactions"]

    print("\n--- Transaction History ---\n")
    data_prsnt=False
    print(f"{'Date':<30}{'Name':<25}{'Price':<15}{'Qty':<10}{'GST':<10}{'Total':<15}")
    print("-" * 110)
    for row in sheet.iter_rows(min_row=2,values_only=True):
        date=row[0]
        name =row[1]
        price = row[2]
        qty =row[3]
        gst=row[4]
        total=row[5]
        data_prsnt=True
        print(f"{date:<30}{name:<25}{price:<15}{qty:<10}{gst:<10}{total:<15}")
    if data_prsnt==False:
        print("\nNO Data Found!")

def admin_menu():
    while True:
        clear_screen()
        show_welcome_screen()
        print("--- ADMIN MENU ---")
        print("1 Add Product")
        print("2 Update Product")
        print("3 Delete Product")
        print("4 View Products")
        print("5 View Transaction History")
        print("6 Generate Bill / Sell Product")
        print("7 Logout")
        print("8 Exit\n")

        choice = input("Enter choice: ")

        if choice == "1":
            add_product()

        elif choice == "2":
            update_product()

        elif choice == "3":
            delete_product()

        elif choice == "4":
            view_products()

        elif choice == "5":
            view_transactions()

        elif choice == "6":
            sell_product()

        elif choice == "7":
            break

        elif choice == "8":
            exit()

        else:
            print("\nInvalid choice.")

        input("\nPress Enter to return to Admin Menu.")

def employee_menu():
    while True:
        clear_screen()
        show_welcome_screen()
        print("\n--- EMPLOYEE MENU ---")
        print("1 View Products")
        print("2 Sell Product")
        print("3 Logout")
        print("4 Exit")

        choice = input("Enter choice: ")

        if choice == "1":
            view_products()

        elif choice == "2":
            sell_product()

        elif choice == "3":
            break

        elif choice == "4":
            exit()
        else:
            print("\nInvalid choice.")

        input("\nPress Enter to return to Employee Menu.")


def main():
    init_inventory()

    while True:
        role = login()

        if role == "admin":
            admin_menu()

        elif role == "employee":
            employee_menu()

if __name__ == "__main__":
    main()